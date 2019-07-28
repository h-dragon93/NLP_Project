from eunjeon import Mecab
from openpyxl import load_workbook
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize, sent_tokenize
import math
import re

load_wb = load_workbook("C:/Users/HanYong/PycharmProjects/BAEKJOON/Naver/naver_news_crawling-master2019-5-6.xlsx",
                        data_only=True)
load_ws = load_wb['sheet1']

text1 = ""
# 셀 좌표로 값 출력
for i in range(2, 9970):
    titles = load_ws.cell(i, 3).value

    contents = load_ws.cell(i, 5).value

    text1 += (titles + contents + "\n")


# print(text1)

def onlyNouns(article):
    result = ""
    mecab = Mecab()
    sample = mecab.pos(article)
    for x in sample:
        if 'NNG' in x or 'NNP' in x or 'NNB' in x or 'NP' in x:
            result += (x[0]) + ' '
            # print(result)
    return result


text1 = text1.split("\n")
# print(text1)
li_noun = []

for i in text1:
    i = onlyNouns(i)
    li_noun.append(i)
# print(li_noun)
text1 = "\n".join(li_noun)


def get_doc(sent):
    """
    이 함수는 텍스트를 문장으로 나누고 각 문장을 document로
    여겨서 각 document의 총 단어 등장 수를 카운트
    """
    doc_info = []
    i = 0
    for x in sent:
        i += 1
        count = count_words(x)
        temp = {'doc_id': i, 'doc_length': count}
        doc_info.append(temp)
    return doc_info


def count_words(sent):
    """
    이 함수는 각 input 텍스트의 총 단어수를 리턴
    """
    count = 0
    words = krword_tokenize(sent)
    # print(words)
    for word in words:
        count += 1
    return count


def krword_tokenize(sent):
    result = []
    mecab = Mecab()
    sample = mecab.pos(sent)
    for x in sample:
        if 'NNG' in x or 'NNP' in x or 'NNB' in x or 'NP' in x:
            result.append(x)
    return result


def create_freq_dict(sents):
    """
    이 함수는 각 사전의 단어들에 대한 빈도수를 dictionary로 만듬
    """
    i = 0
    freqDict_list = []
    for sent in sents:
        i += 1
        freq_dict = {}
        words = word_tokenize(sent)
        for word in words:
            word = word.lower()
            if word in freq_dict:
                freq_dict[word] += 1
            else:
                freq_dict[word] = 1
            temp = {'doc_id': i, 'freq_dict': freq_dict}
        freqDict_list.append(temp)
    return freqDict_list


def computeTF(doc_info, freqDict_list):
    """
    tf = (document 내에서 단어 등장 빈도 / document내의 전체 단어 수)
    """

    TF_scores = []
    for tempDict in freqDict_list:
        id = tempDict['doc_id']
        for k in tempDict['freq_dict']:
            temp = {'doc_id': id,
                    'TF_score': tempDict['freq_dict'][k] / doc_info[id - 1]['doc_length'],
                    'key': k}
            TF_scores.append(temp)
    return TF_scores


def computeIDF(doc_info, freqDict_list):
    """
    idf = log(전체 문서 수/ 단어가 드 안에 들어있는 문서 수)
    """
    IDF_scores = []
    counter = 0
    for dict in freqDict_list:
        counter += 1
        for k in dict['freq_dict'].keys():
            count = sum([k in tempDict['freq_dict'] for tempDict in freqDict_list])
            temp = {'doc_id': counter, 'IDF_score': math.log(len(doc_info) / count), 'key': k}

            IDF_scores.append(temp)

    return IDF_scores


def computeTFIDF(TF_scores, IDF_scores):
    TFIDF_scores = []
    for j in IDF_scores:
        for i in TF_scores:
            if j['key'] == i['key'] and j['doc_id'] == i['doc_id']:
                temp = {'doc_id': j['doc_id'],
                        'TFIDF_score': j['IDF_score'] * i['TF_score'],
                        'key': i['key']}
        TFIDF_scores.append(temp)
    return TFIDF_scores


def remove_string_special_characters(s):
    """
    s : 문자열
    """
    stripped = re.sub('[^\w\s]', '', s)
    # 문자열 중간 공백 삭제
    stripped = re.sub('\s+', ' ', stripped)
    # 처음과 끝 공백 삭제
    stripped = stripped.strip()

    return stripped

text_sents = text1.split("\n")
doc_info = get_doc(text_sents)
#print(doc_info)

freqDict_list = create_freq_dict(text_sents)
#print(freqDict_list)
TF_scores = computeTF(doc_info, freqDict_list)
IDF_scores = computeIDF(doc_info, freqDict_list)

#doc_info

#print("TF_Score", TF_scores)

#print("IDF_SCore", IDF_scores)

TFIDF_scores = computeTFIDF(TF_scores, IDF_scores)
#print("TF_IDF Score", TFIDF_scores)

li = []
for i in TFIDF_scores :
    li.append(i['TFIDF_score'])
li.sort()
li.reverse()
li = li[0:100]
print(li)

for i in TFIDF_scores :
    if i['TFIDF_score'] in li :
        print(i['key'])


