from eunjeon import Mecab
from openpyxl import load_workbook
import re
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.preprocessing import normalize
import numpy as np

load_wb = load_workbook("C:/Users/HanYong/PycharmProjects/BAEKJOON/Naver/naver_news_crawling-master2019-5-6.xlsx",
                        data_only=True)
load_ws = load_wb['sheet1']


def cleanText(rawText):
    toBeTerminated = '^[ㄱ-힣]'
    strippedText = re.sub(toBeTerminated, '', rawText)
    return strippedText


wordList = []

# 셀 좌표로 값 출력
for i in range(1, 300):

    mecab = Mecab()
    titles = load_ws.cell(i, 3).value
    # titles = cleanText(titles)

    contents = load_ws.cell(i, 5).value
    # contents = cleanText(contents)

    texts = titles + contents

    # print(texts)
    malist = mecab.pos(texts)
    nouns = []
    for x in malist:
        if 'NNG' in x or 'NNP' in x or 'NNB' in x or 'NP' in x:
            keyword = x[0]
            nouns.append(keyword)
    # print(nouns)
    wordList.append(' '.join(nouns))

print(wordList[1:])

class GraphMatrix(object):
    def __init__(self):
        self.tfidf = TfidfVectorizer()
        self.cnt_vec = CountVectorizer()
        self.graph_sentence = []

    def build_sent_graph(self, sentence):
        tfidf_mat = self.tfidf.fit_transform(sentence).toarray()
        self.graph_sentence = np.dot(tfidf_mat, tfidf_mat.T)
        return self.graph_sentence

    def build_words_graph(self, sentence):
        cnt_vec_mat = normalize(self.cnt_vec.fit_transform(sentence).toarray().astype(float), axis=0)
        vocab = self.cnt_vec.vocabulary_
        return np.dot(cnt_vec_mat.T, cnt_vec_mat), {vocab[word]: word for word in vocab}

class Rank(object):
    def get_ranks(self, graph, d=0.85):  # d = damping factor
        A = graph
        matrix_size = A.shape[0]
        for id in range(matrix_size):
            A[id, id] = 0  # diagonal 부분을 0으로
            link_sum = np.sum(A[:, id])  # A[:, id] = A[:][id]
            if link_sum != 0:
                A[:, id] /= link_sum
            A[:, id] *= -d
            A[id, id] = 1

        B = (1 - d) * np.ones((matrix_size, 1))
        ranks = np.linalg.solve(A, B)  # 연립방정식 Ax = b
        return {idx: r[0] for idx, r in enumerate(ranks)}

graph_matrix = GraphMatrix()
sent_graph = graph_matrix.build_sent_graph(wordList)
words_graph, idx2word = graph_matrix.build_words_graph(wordList)

rank = Rank()
sent_rank_idx = rank.get_ranks(sent_graph)
sorted_sent_rank_idx = sorted(sent_rank_idx, key=lambda k: sent_rank_idx[k], reverse=True)

word_rank_idx = rank.get_ranks(words_graph)
sorted_word_rank_idx = sorted(word_rank_idx, key=lambda k: word_rank_idx[k], reverse=True)

##summarize
sent_num = 10
summary = []
index = []

for idx in sorted_sent_rank_idx[:sent_num]:
    index.append(idx)
index.sort()

for idx in index:
    summary.append(wordList[idx])
print(summary)

word_num = 10
rank = Rank()
rank_idx = rank.get_ranks(words_graph)
sorted_rank_idx = sorted(rank_idx, key=lambda k: rank_idx[k], reverse=True)
keywords = []
index = []
for idx in sorted_rank_idx[:word_num]:
    index.append(idx)

for idx in index:
    keywords.append(idx2word[idx])

print(keywords)
